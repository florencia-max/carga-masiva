"""
AllRide – Cuadratura de viajes v2.0
Flujo completo en 7 pasos:
  1. Paradas — verificar acceso y detectar faltantes
  2. Rutas regulares — detectar rutas nuevas a crear
  3. Edición horarios regulares — ajustar hora en viajes existentes
  4. Edición horarios SPOT — ajustar hora en viajes SPOT existentes
  5. Creación viajes regulares — one-time services por empresa
  6. Creación viajes SPOT — ODD por empresa
  7. Cancelación masiva — marcar con X lo que sobra
"""

import streamlit as st
import pandas as pd
import re
import zipfile
from io import BytesIO
from datetime import timedelta, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="AllRide – Cuadratura", page_icon="🚌", layout="wide")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
EMP_NORM = {
    "KAUFMANN":"KAUFMANN",
    "FALABELLA - LA SERENA":"Falabella - La Serena",
    "FALABELLA - VALPARAÍSO":"Falabella - Valparaíso",
    "FALABELLA - VALPARAISO":"Falabella - Valparaíso",
    "TRADIS LOGÍSTICA FALABELLA":"TRADIS Logística Falabella",
    "TRADIS LOGISTICA FALABELLA":"TRADIS Logística Falabella",
    "TÍO TOMATE":"Tío Tomate","TIO TOMATE":"Tío Tomate",
    "FALABELLA - RM":"Falabella - RM",
    "FALABELLA - RANCAGUA":"Falabella - Rancagua",
    "FALABELLA - ANTOFAGASTA":"Falabella - Antofagasta",
    "FALABELLA - IQUIQUE":"Falabella - Iquique",
    "FALABELLA - VALDIVIA":"Falabella - Valdivia",
    "FALABELLA - TALCA":"Falabella - Talca",
    "SODIMAC - CORONEL":"Sodimac - Coronel",
    "GRUPO ALCANSA":"Grupo Alcansa","ANDES MOTOR":"Andes Motor",
}
TIPO_AR = {
    "SALIDA CALENDARIZADA":"REGULAR","SALIDA REALIZADA":"REGULAR",
    "SERVICIO REGULAR":"REGULAR","SERVICIO ESPECIAL":"SPOT",
}
PARADA_FALLBACK = "Parada TRP"
PARADA_INGRESO  = "Parada"

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def norm(s):
    s = str(s).strip().upper()
    for a,b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
        s = s.replace(a, b)
    return s

def norm_empresa(s):
    n = norm(str(s))
    return EMP_NORM.get(n, str(s).strip())

def norm_ruta(s):
    s = re.sub(r"^RDD\s*-\s*", "", str(s).strip(), flags=re.IGNORECASE)
    return norm(s)

def is_spot(ruta):
    return "SPOT" in norm(str(ruta))

def parse_hora(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return "00:00"
    if isinstance(val, datetime): return val.strftime("%H:%M")
    if isinstance(val, timedelta):
        t = int(val.total_seconds())
        return f"{t//3600:02d}:{(t%3600)//60:02d}"
    if isinstance(val, (int, float)):
        frac = val - int(val)
        tm = round(frac * 1440)
        return f"{(tm//60)%24:02d}:{tm%60:02d}"
    s = str(val).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else "00:00"

def sumar15(h):
    m = re.match(r"^(\d{1,2}):(\d{2})$", h)
    if not m: return h
    t = int(m.group(1))*60 + int(m.group(2)) + 15
    t %= 1440
    return f"{t//60:02d}:{t%60:02d}"

def h2min(h):
    m = re.match(r"^(\d{1,2}):(\d{2})$", h)
    return int(m.group(1))*60 + int(m.group(2)) if m else 0

def parse_fecha_ar(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return ""
    if isinstance(val, datetime): return val.strftime("%d/%m/%Y")
    s = str(val).split(",")[0].strip()
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", s)
    return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}" if m else s

def parse_hora_ar(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return "00:00"
    if isinstance(val, datetime): return val.strftime("%H:%M")
    parts = str(val).split(",")
    return parse_hora(parts[1].strip()) if len(parts) > 1 else parse_hora(val)

def fmt_fecha_excel(fecha_str):
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", str(fecha_str))
    if m:
        try: return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
        except: pass
    return fecha_str

def detectar_col(cols, candidates):
    cn = {norm(c): c for c in cols}
    for c in candidates:
        if norm(c) in cn: return cn[norm(c)]
    return None

def extraer_ultima_parada(recorrido_str):
    if not recorrido_str or str(recorrido_str).strip() in ["", "nan", "NaN", "None"]:
        return None
    s = str(recorrido_str).strip()
    for sep in ["\n", ";", "|", " → ", " -> "]:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if parts: return parts[-1]
    lines = [l.strip() for l in s.splitlines() if l.strip()]
    if len(lines) > 1: return lines[-1]
    # Último elemento separado por " - " solo si hay múltiples
    if " - " in s:
        parts = [p.strip() for p in s.split(" - ") if p.strip()]
        if len(parts) > 1: return parts[-1]
    return None

def stops_empresa(stops_df, empresa):
    result = set()
    en = norm(empresa)
    for _, row in stops_df.iterrows():
        comunidades = str(row.get("Comunidades", ""))
        partes = [norm(p.strip()) for p in re.split(r"[,;]", comunidades)]
        if en in partes:
            result.add(str(row["Nombre parada"]).strip())
    return result

def gen_zip(archivos):
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, data in archivos.items():
            zf.writestr(nombre, data)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def leer_consolidado(f):
    xl = pd.ExcelFile(f)
    hoja = None
    for h in xl.sheet_names:
        if norm(h) in ["RESUMEN","PROGRAMACION RM","PROGRAMACION","PROGRAMA","PROGRAMACIÓN RM"]:
            hoja = h; break
    if not hoja: hoja = xl.sheet_names[0]
    return pd.read_excel(f, sheet_name=hoja), hoja

def procesar_consolidado(df):
    cols = list(df.columns)
    col_ruta      = detectar_col(cols, ["NUEVO NOMBRE RUTA FINAL","NOMBRE RUTA FINAL","NOMBRE RUTA"])
    col_postura   = detectar_col(cols, ["HORA DE POSTURA"])
    col_hora      = detectar_col(cols, ["HORA DE LLEGADA A BODEGA","HORA DE LLEGADA\n HORA DE SALIDA","HORA DE SALIDA","HORA DE LLEGADA","HORA"])
    col_fecha     = detectar_col(cols, ["FECHA"])
    col_empresa   = detectar_col(cols, ["EMPRESA"])
    col_tipo      = detectar_col(cols, ["TIPO DE PEDIDO","TIPO"])
    col_bodega    = detectar_col(cols, ["BODEGA"])
    col_recorrido = detectar_col(cols, ["RECORRIDO"])
    col_tipo_mov  = detectar_col(cols, ["TIPO ","TIPO"])

    if not col_ruta or not col_fecha:
        st.error("❌ No se encontró columna de ruta o fecha en el consolidado.")
        st.stop()

    out = pd.DataFrame()
    out["ruta_orig"]    = df[col_ruta].astype(str).str.strip()
    out["ruta_norm"]    = out["ruta_orig"].apply(norm_ruta)
    out["empresa_orig"] = df[col_empresa].astype(str).str.strip() if col_empresa else "SIN EMPRESA"
    out["empresa_norm"] = out["empresa_orig"].apply(norm_empresa)
    out["fecha"]        = pd.to_datetime(df[col_fecha], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
    out["tipo_norm"]    = df[col_tipo].apply(lambda x: norm(str(x))) if col_tipo else "REGULAR"
    out["bodega"]       = df[col_bodega].astype(str).apply(lambda x: x.split("\n")[0].strip()) if col_bodega else ""
    out["recorrido"]    = df[col_recorrido].astype(str) if col_recorrido else ""
    out["tipo_mov"]     = df[col_tipo_mov].astype(str).str.strip().str.upper() if col_tipo_mov else ""

    if col_postura:
        out["hora_postura"] = df[col_postura].apply(parse_hora)
        out["hora_allride"] = out["hora_postura"].apply(sumar15)
    else:
        out["hora_postura"] = df[col_hora].apply(parse_hora) if col_hora else "00:00"
        out["hora_allride"] = out["hora_postura"]

    out["es_spot"] = out["ruta_orig"].apply(is_spot)
    out = out[out["ruta_orig"].notna() & ~out["ruta_orig"].isin(["","NAN","nan","None"])]
    out = out[out["fecha"].notna() & ~out["fecha"].isin(["","NaT","NaN"])]
    return out.reset_index(drop=True)

def procesar_allride(df):
    cols = list(df.columns)
    col_fecha_raw = detectar_col(cols, ["Fecha estimada de inicio","Fecha de inicio"])
    col_ruta      = detectar_col(cols, ["Ruta"])
    col_empresa   = detectar_col(cols, ["Comunidades"])
    col_tipo      = detectar_col(cols, ["Tipo"])
    col_id        = detectar_col(cols, ["ID de servicio","Servicio","ID"])
    col_estado    = detectar_col(cols, ["Estado"])

    fecha_raw = df[col_fecha_raw] if col_fecha_raw else pd.Series([""] * len(df))
    out = pd.DataFrame()
    out["id"]           = df[col_id].astype(str).str.strip() if col_id else ""
    out["ruta_orig"]    = df[col_ruta].astype(str).str.strip() if col_ruta else ""
    out["ruta_norm"]    = out["ruta_orig"].apply(norm_ruta)
    out["empresa_orig"] = df[col_empresa].astype(str).str.strip() if col_empresa else ""
    out["empresa_norm"] = out["empresa_orig"].apply(norm_empresa)
    out["tipo_orig"]    = df[col_tipo].astype(str).str.strip() if col_tipo else ""
    out["tipo_norm"]    = out["tipo_orig"].apply(lambda x: TIPO_AR.get(norm(x), norm(x)))
    out["estado"]       = df[col_estado].astype(str).str.strip() if col_estado else ""
    out["fecha"]        = fecha_raw.apply(parse_fecha_ar)
    out["hora"]         = fecha_raw.apply(parse_hora_ar)
    out["es_spot"]      = out["ruta_norm"].apply(is_spot)
    return out.reset_index(drop=True)

def cuadrar(cli, ar):
    ar_grupos = {}
    for idx, row in ar.iterrows():
        k = (row["fecha"], row["ruta_norm"], row["empresa_norm"])
        ar_grupos.setdefault(k, []).append((idx, row))

    ar_usados = set()
    res = {"ok":[], "editar":[], "cancelar":[], "crear":[]}

    for _, crow in cli.iterrows():
        k = (crow["fecha"], crow["ruta_norm"], crow["empresa_norm"])
        hora_obj = crow["hora_allride"]
        candidatos = ar_grupos.get(k, [])

        if not candidatos:
            res["crear"].append(crow); continue

        exactos = [(i,r) for i,r in candidatos if r["hora"] == hora_obj]
        if exactos:
            idx_m, row_m = exactos[0]
            ar_usados.add(idx_m)
            res["ok"].append(row_m); continue

        no_usados = [(i,r) for i,r in candidatos if i not in ar_usados]
        if not no_usados:
            res["crear"].append(crow); continue

        no_usados.sort(key=lambda x: abs(h2min(x[1]["hora"]) - h2min(hora_obj)))
        idx_e, row_e = no_usados[0]
        ar_usados.add(idx_e)
        res["editar"].append({
            "ar_idx": idx_e, "ar_row": row_e, "cli_row": crow,
            "hora_nueva": hora_obj, "hora_actual": row_e["hora"]
        })

    for k, grupo in ar_grupos.items():
        for idx, row in grupo:
            if idx not in ar_usados:
                res["cancelar"].append(row)

    return res

# ══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS PARADAS Y RUTAS
# ══════════════════════════════════════════════════════════════════════════════
def analizar_paradas(cli, stops_df, routes_df):
    stops_norm = {}
    for _, r in stops_df.iterrows():
        n = norm(str(r["Nombre parada"]))
        stops_norm[n] = {"nombre": str(r["Nombre parada"]), "comunidades": str(r.get("Comunidades",""))}

    routes_norm = {}
    for _, r in routes_df.iterrows():
        n = norm_ruta(str(r.get("Nombre ruta","")))
        routes_norm[n] = r

    adv_acceso, paradas_faltantes = [], []
    vistas = set()

    for _, crow in cli.iterrows():
        if crow["es_spot"]: continue
        empresa = crow["empresa_norm"]
        ruta_n  = crow["ruta_norm"]
        ruta_ar = routes_norm.get(ruta_n)

        paradas = []
        if ruta_ar is not None:
            po  = str(ruta_ar.get("Nombre Parada Origen","")).strip()
            pd_ = str(ruta_ar.get("Nombre Parada Destino","")).strip()
            if po  and po  not in ["nan",""]: paradas.append(po)
            if pd_ and pd_ not in ["nan",""]: paradas.append(pd_)
        else:
            if crow.get("bodega"): paradas.append(crow["bodega"])

        for parada in paradas:
            pk = norm(parada)
            key = f"{pk}|{norm(empresa)}"
            if key in vistas: continue
            vistas.add(key)

            if pk not in stops_norm:
                if not any(p["Nombre parada"] == parada for p in paradas_faltantes):
                    paradas_faltantes.append({"Nombre parada": parada, "Empresa": empresa, "Ruta": crow["ruta_orig"]})
            else:
                info = stops_norm[pk]
                comps = [norm(c.strip()) for c in re.split(r"[,;]", info["comunidades"])]
                if norm(empresa) not in comps:
                    adv_acceso.append({
                        "Parada": info["nombre"],
                        "Empresa que la necesita": empresa,
                        "Comunidades actuales": info["comunidades"],
                        "Ruta": crow["ruta_orig"]
                    })

    return adv_acceso, paradas_faltantes

def analizar_rutas(cli, routes_df):
    routes_norm = set(norm_ruta(str(r["Nombre ruta"])) for _, r in routes_df.iterrows())
    nuevas = {}
    for _, crow in cli.iterrows():
        if crow["es_spot"]: continue
        rn = crow["ruta_norm"]
        if rn not in routes_norm and rn not in nuevas:
            nuevas[rn] = {
                "Nombre ruta": crow["ruta_orig"],
                "Empresa": crow["empresa_norm"],
                "Tipo mov": crow.get("tipo_mov",""),
                "Bodega": crow.get("bodega","")
            }
    return list(nuevas.values())

# ══════════════════════════════════════════════════════════════════════════════
# GENERADORES
# ══════════════════════════════════════════════════════════════════════════════
def gen_stops_creation(paradas, tmpl_bytes):
    wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
    ws = wb.active
    font = Font(name="Calibri", size=11)
    for i, p in enumerate(paradas, 2):
        ws.cell(i,1, p["Nombre parada"]).font = font
        ws.cell(i,2, "").font = font   # Lat vacío
        ws.cell(i,3, "").font = font   # Lon vacío
        ws.cell(i,4, p["Empresa"]).font = font
        ws.cell(i,5, "Sí").font = font
        ws.cell(i,6, p["Nombre parada"]).font = font
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

def gen_routes_creation(rutas, tmpl_bytes):
    wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row: cell.value = None
    font = Font(name="Calibri", size=11)
    for i, r in enumerate(rutas, 2):
        tipo = str(r.get("Tipo mov","")).upper()
        bodega = str(r.get("Bodega","")).strip()
        if "SALIDA" in tipo:
            origen  = bodega or "Parada"
            destino = PARADA_FALLBACK
        else:
            origen  = PARADA_INGRESO
            destino = bodega or "Parada"
        ws.cell(i, 1, r["Nombre ruta"]).font = font
        ws.cell(i, 2, r["Nombre ruta"]).font = font
        ws.cell(i, 3, "-").font = font
        ws.cell(i, 4, "Ida").font = font
        ws.cell(i, 5, origen).font = font
        ws.cell(i, 6, destino).font = font
        ws.cell(i, 7, "-").font = font
        ws.cell(i, 19, "Si").font = font
        ws.cell(i, 22, r["Empresa"]).font = font
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

def gen_edicion_horarios(editar_filas, tmpl_bytes):
    font = Font(name="Calibri", size=12)
    if not tmpl_bytes:
        rows = [{"ID de servicio":e["ar_row"]["id"],"Hora del servicio":e["hora_nueva"],
                 "Ruta":e["ar_row"]["ruta_orig"],"Empresa":e["ar_row"]["empresa_orig"],
                 "Fecha":e["ar_row"]["fecha"],"Hora actual":e["hora_actual"]} for e in editar_filas]
        buf = BytesIO(); pd.DataFrame(rows).to_excel(buf, index=False); return buf.getvalue()

    wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
    ws = wb.active
    header = [ws.cell(1,c).value for c in range(1, ws.max_column+2)]
    def fc(names):
        for n in names:
            for i,h in enumerate(header):
                if h and norm(str(h)) == norm(n): return i+1
        return None
    col_id   = fc(["ID de servicio","ID"])
    col_hora = fc(["Hora del servicio","Hora de inicio","Hora"])
    col_ruta = fc(["Ruta","Nombre ruta"])
    ws.delete_rows(2, ws.max_row)
    for i, e in enumerate(editar_filas, 2):
        ar = e["ar_row"]
        if col_id:   ws.cell(i, col_id,   ar["id"]).font = font
        if col_hora: ws.cell(i, col_hora, e["hora_nueva"]).font = font
        if col_ruta: ws.cell(i, col_ruta, ar["ruta_orig"]).font = font
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

def gen_one_time_services(crear_filas, tmpl_bytes):
    por_emp = {}
    for row in crear_filas:
        if row["es_spot"]: continue
        por_emp.setdefault(row["empresa_norm"], []).append(row)

    archivos = {}
    for emp, rows in por_emp.items():
        wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
        ws = wb.active
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in r: c.value = None
        font = Font(name="Calibri", size=11)
        for i, row in enumerate(rows, 2):
            ws.cell(i,1, row["ruta_orig"]).font = font
            fecha = fmt_fecha_excel(row["fecha"])
            ws.cell(i,2, fecha).font = font
            ws.cell(i,2).number_format = "DD-MM-YYYY"
            ws.cell(i,3, row["hora_allride"]).font = font
            ws.cell(i,4, 1).font = font
        buf = BytesIO(); wb.save(buf)
        nombre = re.sub(r"[^\w\-]", "_", emp)[:40]
        archivos[f"one_time_services_{nombre}.xlsx"] = buf.getvalue()
    return archivos

def gen_odd_spots(crear_filas, tmpl_bytes, stops_df):
    por_emp = {}
    for row in crear_filas:
        if not row["es_spot"]: continue
        por_emp.setdefault(row["empresa_norm"], []).append(row)

    stops_cache = {}
    archivos = {}

    for emp, rows in por_emp.items():
        if emp not in stops_cache:
            stops_cache[emp] = stops_empresa(stops_df, emp)
        emp_stops = stops_cache[emp]

        wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
        ws = wb.active
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in r: c.value = None
        font = Font(name="Calibri", size=11)

        for i, row in enumerate(rows, 2):
            tipo_mov  = str(row.get("tipo_mov","")).upper()
            bodega    = str(row.get("bodega","")).strip()
            recorrido = str(row.get("recorrido","")).strip()
            es_salida = "SALIDA" in tipo_mov

            if es_salida:
                npo = bodega if bodega and bodega in emp_stops else PARADA_FALLBACK
                ultima = extraer_ultima_parada(recorrido)
                npd = ultima if ultima and ultima in emp_stops else PARADA_FALLBACK
                parada_usuario = ultima or bodega or ""
                od = "Origen"
            else:
                ultima = extraer_ultima_parada(recorrido)
                parada_usuario = ultima or bodega or ""
                npo = None
                npd = bodega if bodega and bodega in emp_stops else PARADA_INGRESO
                od = "Destino"

            ws.cell(i,  1, i-1).font = font
            ws.cell(i,  2, row["ruta_orig"]).font = font
            ws.cell(i,  3, "Servicios Dinámicos").font = font
            ws.cell(i,  4, od).font = font
            ws.cell(i,  5, row["fecha"]).font = font
            ws.cell(i,  6, row["hora_allride"]).font = font
            ws.cell(i,  7, npo or "").font = font
            ws.cell(i,  8, npd or "").font = font
            ws.cell(i,  9, parada_usuario).font = font
            ws.cell(i, 10, 1).font = font
            ws.cell(i, 11, "").font = font
            ws.cell(i, 12, "").font = font
            ws.cell(i, 13, "").font = font
            ws.cell(i, 14, emp).font = font

        buf = BytesIO(); wb.save(buf)
        nombre = re.sub(r"[^\w\-]", "_", emp)[:40]
        archivos[f"ODD_SPOTS_{nombre}.xlsx"] = buf.getvalue()
    return archivos

def gen_cancelacion(cancelar_filas, tmpl_bytes):
    ids = set(str(r["id"]) for r in cancelar_filas)
    wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
    ws = wb.active
    header = [ws.cell(1,c).value for c in range(1, ws.max_column+2)]
    try: col_id = header.index("ID de servicio")+1
    except: col_id = 1
    try: col_cancel = header.index("Cancelar")+1
    except: col_cancel = ws.max_column
    fx   = Font(name="Calibri", size=12, bold=True, color="C0392B")
    fill = PatternFill("solid", fgColor="FDEDEC")
    for r in range(2, ws.max_row+1):
        id_val = str(ws.cell(r, col_id).value or "").strip()
        if id_val in ids:
            c = ws.cell(r, col_cancel)
            c.value="X"; c.font=fx; c.fill=fill
            c.alignment=Alignment(horizontal="center")
        else:
            ws.cell(r, col_cancel).value = None
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# UI — SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
st.title("🚌 AllRide – Cuadratura de viajes")
st.caption("Flujo: 1 Paradas → 2 Rutas → 3-4 Editar → 5-6 Crear → 7 Cancelar")

with st.sidebar:
    st.header("📁 Obligatorios")
    f_consolidado   = st.file_uploader("Consolidado del cliente", type=["xlsx","xls"])
    f_allride       = st.file_uploader("Exportación AllRide", type=["xlsx","xls"])
    f_cancel_tmpl   = st.file_uploader("Plantilla cancelación masiva", type=["xlsx","xls"])

    st.header("📁 Referencia AllRide")
    f_stops_list    = st.file_uploader("Paradas existentes (stops list)", type=["xlsx","xls"])
    f_routes_list   = st.file_uploader("Rutas existentes (routes edition)", type=["xlsx","xls"])

    st.header("📁 Templates creación")
    f_ots_tmpl      = st.file_uploader("Template One Time Services", type=["xlsx","xls"])
    f_odd_tmpl      = st.file_uploader("Template ODD Spots", type=["xlsx","xls"])
    f_routes_tmpl   = st.file_uploader("Template creación rutas", type=["xlsx","xls"])
    f_stops_tmpl    = st.file_uploader("Template creación paradas", type=["xlsx","xls"])

    st.header("📁 Templates edición (opcional)")
    f_edit_reg_tmpl  = st.file_uploader("Template edición horarios reg.", type=["xlsx","xls"])
    f_edit_spot_tmpl = st.file_uploader("Template edición horarios SPOT", type=["xlsx","xls"])

if not all([f_consolidado, f_allride, f_cancel_tmpl]):
    st.info("⬅️ Sube al menos los 3 archivos obligatorios para comenzar.")
    st.stop()

# ── PROCESAR ─────────────────────────────────────────────────────────────────
with st.spinner("Procesando..."):
    cli_raw, hoja_cli = leer_consolidado(f_consolidado)
    ar_raw  = pd.read_excel(f_allride)
    stops_df  = pd.read_excel(f_stops_list)  if f_stops_list  else pd.DataFrame(columns=["Comunidades","Nombre parada","Lat","Lon"])
    routes_df = pd.read_excel(f_routes_list) if f_routes_list else pd.DataFrame(columns=["Nombre ruta","Nombre Parada Origen","Nombre Parada Destino","Comunidades (Nombre exacto, separados por comas)"])

    cli = procesar_consolidado(cli_raw)
    ar  = procesar_allride(ar_raw)
    res = cuadrar(cli, ar)

    editar_reg  = [e for e in res["editar"] if not e["ar_row"]["es_spot"]]
    editar_spot = [e for e in res["editar"] if  e["ar_row"]["es_spot"]]
    crear_reg   = [r for r in res["crear"]  if not r["es_spot"]]
    crear_spot  = [r for r in res["crear"]  if  r["es_spot"]]

st.caption(f"📋 Consolidado: hoja **{hoja_cli}** · {len(cli)} viajes cliente · {len(ar)} viajes AllRide")

# ── MÉTRICAS ─────────────────────────────────────────────────────────────────
st.divider()
c = st.columns(7)
c[0].metric("✅ OK",           len(res["ok"]))
c[1].metric("✏️ Editar reg.",  len(editar_reg))
c[2].metric("✏️ Editar SPOT",  len(editar_spot))
c[3].metric("➕ Crear reg.",   len(crear_reg))
c[4].metric("🔶 Crear SPOT",   len(crear_spot))
c[5].metric("❌ Cancelar",     len(res["cancelar"]))
c[6].metric("📋 Total cliente",len(cli))

# ── TABS ─────────────────────────────────────────────────────────────────────
tabs = st.tabs([
    "1️⃣ Paradas",
    "2️⃣ Rutas nuevas",
    "3️⃣ Editar regulares",
    "4️⃣ Editar SPOT",
    "5️⃣ Crear regulares",
    "6️⃣ Crear SPOT",
    "7️⃣ Cancelación",
    "📦 Todo en ZIP",
])

# ─── TAB 1 ───────────────────────────────────────────────────────────────────
with tabs[0]:
    st.subheader("1️⃣ Verificación de paradas")
    if not f_stops_list:
        st.warning("⚠️ Sube el archivo de paradas existentes para este análisis.")
    else:
        adv_acceso, paradas_faltantes = analizar_paradas(cli, stops_df, routes_df)

        if adv_acceso:
            st.warning(f"⚠️ **{len(adv_acceso)} paradas sin acceso para el cliente** — revisar manualmente en AllRide:")
            st.dataframe(pd.DataFrame(adv_acceso), use_container_width=True)
        else:
            st.success("✅ Sin problemas de acceso a paradas existentes.")

        if paradas_faltantes:
            st.error(f"🔴 **{len(paradas_faltantes)} paradas no existen en AllRide** — crearlas primero:")
            st.dataframe(pd.DataFrame(paradas_faltantes), use_container_width=True)
            if f_stops_tmpl:
                data = gen_stops_creation(paradas_faltantes, f_stops_tmpl.read())
                st.download_button("⬇️ Descargar template paradas a crear",
                    data, "paradas_a_crear.xlsx", use_container_width=True)
                st.caption("⚠️ Lat/Lon deben completarse manualmente antes de subir a AllRide.")
            else:
                st.info("Sube el template de creación de paradas para generar el archivo.")
        else:
            st.success("✅ Todas las paradas necesarias existen en AllRide.")

# ─── TAB 2 ───────────────────────────────────────────────────────────────────
with tabs[1]:
    st.subheader("2️⃣ Rutas regulares a crear")
    if not f_routes_list:
        st.warning("⚠️ Sube el archivo de rutas existentes para este análisis.")
    else:
        rutas_nuevas = analizar_rutas(cli, routes_df)
        if rutas_nuevas:
            st.error(f"🔴 **{len(rutas_nuevas)} rutas regulares no existen en AllRide:**")
            st.dataframe(pd.DataFrame(rutas_nuevas), use_container_width=True)
            if f_routes_tmpl:
                data = gen_routes_creation(rutas_nuevas, f_routes_tmpl.read())
                st.download_button("⬇️ Descargar template rutas a crear",
                    data, "rutas_a_crear.xlsx", use_container_width=True)
            else:
                st.info("Sube el template de creación de rutas para generar el archivo.")
        else:
            st.success("✅ Todas las rutas regulares ya existen en AllRide.")

# ─── TAB 3 ───────────────────────────────────────────────────────────────────
with tabs[2]:
    st.subheader("3️⃣ Edición de horarios — regulares")
    if not editar_reg:
        st.success("✅ No hay viajes regulares que editar.")
    else:
        st.dataframe(pd.DataFrame([{
            "Fecha":e["ar_row"]["fecha"],"Ruta":e["ar_row"]["ruta_orig"],
            "Empresa":e["ar_row"]["empresa_orig"],"ID":e["ar_row"]["id"],
            "Hora actual":e["hora_actual"],"→ Hora nueva":e["hora_nueva"],
            "Postura cliente":e["cli_row"].get("hora_postura",""),
        } for e in editar_reg]), use_container_width=True)
        tmpl = f_edit_reg_tmpl.read() if f_edit_reg_tmpl else None
        st.download_button(f"⬇️ Descargar edición horarios regulares ({len(editar_reg)})",
            gen_edicion_horarios(editar_reg, tmpl),
            "Edicion_horarios_regulares.xlsx", use_container_width=True)

# ─── TAB 4 ───────────────────────────────────────────────────────────────────
with tabs[3]:
    st.subheader("4️⃣ Edición de horarios — SPOT")
    if not editar_spot:
        st.success("✅ No hay viajes SPOT que editar.")
    else:
        st.dataframe(pd.DataFrame([{
            "Fecha":e["ar_row"]["fecha"],"Ruta":e["ar_row"]["ruta_orig"],
            "Empresa":e["ar_row"]["empresa_orig"],"ID":e["ar_row"]["id"],
            "Hora actual":e["hora_actual"],"→ Hora nueva":e["hora_nueva"],
            "Postura cliente":e["cli_row"].get("hora_postura",""),
        } for e in editar_spot]), use_container_width=True)
        tmpl = f_edit_spot_tmpl.read() if f_edit_spot_tmpl else None
        st.download_button(f"⬇️ Descargar edición horarios SPOT ({len(editar_spot)})",
            gen_edicion_horarios(editar_spot, tmpl),
            "Edicion_horarios_SPOT.xlsx", use_container_width=True)

# ─── TAB 5 ───────────────────────────────────────────────────────────────────
with tabs[4]:
    st.subheader("5️⃣ Crear viajes regulares — One Time Services")
    if not crear_reg:
        st.success("✅ No hay viajes regulares que crear.")
    else:
        por_emp = {}
        for r in crear_reg: por_emp.setdefault(r["empresa_norm"],[]).append(r)
        st.info(f"**{len(crear_reg)} viajes** en **{len(por_emp)} empresas** — un archivo por empresa.")
        for emp, rows in sorted(por_emp.items()):
            with st.expander(f"**{emp}** — {len(rows)} viajes"):
                st.dataframe(pd.DataFrame([{
                    "Fecha":r["fecha"],"Hora postura":r["hora_postura"],
                    "Hora AllRide":r["hora_allride"],"Ruta":r["ruta_orig"]
                } for r in rows]))
        if f_ots_tmpl:
            tmpl_bytes = f_ots_tmpl.read()
            arch = gen_one_time_services(crear_reg, tmpl_bytes)
            if len(arch) == 1:
                name, data = list(arch.items())[0]
                st.download_button(f"⬇️ {name}", data, name, use_container_width=True)
            else:
                st.download_button(f"⬇️ Descargar todos ({len(arch)} archivos .zip)",
                    gen_zip(arch), "one_time_services.zip",
                    mime="application/zip", use_container_width=True)
        else:
            st.warning("Sube el template de One Time Services para generar los archivos.")

# ─── TAB 6 ───────────────────────────────────────────────────────────────────
with tabs[5]:
    st.subheader("6️⃣ Crear viajes SPOT — ODD")
    if not crear_spot:
        st.success("✅ No hay viajes SPOT que crear.")
    else:
        por_emp = {}
        for r in crear_spot: por_emp.setdefault(r["empresa_norm"],[]).append(r)
        st.info(f"**{len(crear_spot)} viajes SPOT** en **{len(por_emp)} empresas** — un archivo por empresa.")
        for emp, rows in sorted(por_emp.items()):
            with st.expander(f"**{emp}** — {len(rows)} viajes"):
                st.dataframe(pd.DataFrame([{
                    "Fecha":r["fecha"],"Hora":r["hora_allride"],
                    "Ruta":r["ruta_orig"],"Tipo":r.get("tipo_mov","")
                } for r in rows]))
        if f_odd_tmpl and f_stops_list:
            arch = gen_odd_spots(crear_spot, f_odd_tmpl.read(), stops_df)
            if len(arch) == 1:
                name, data = list(arch.items())[0]
                st.download_button(f"⬇️ {name}", data, name, use_container_width=True)
            else:
                st.download_button(f"⬇️ Descargar todos ({len(arch)} archivos .zip)",
                    gen_zip(arch), "ODD_SPOTS.zip",
                    mime="application/zip", use_container_width=True)
        else:
            if not f_odd_tmpl:   st.warning("Sube el template ODD Spots.")
            if not f_stops_list: st.warning("Sube el archivo de paradas existentes.")

# ─── TAB 7 ───────────────────────────────────────────────────────────────────
with tabs[6]:
    st.subheader("7️⃣ Cancelación masiva")
    if not res["cancelar"]:
        st.success("✅ No hay viajes que cancelar.")
    else:
        st.dataframe(pd.DataFrame([{
            "Fecha":r["fecha"],"Hora":r["hora"],"Ruta":r["ruta_orig"],
            "Empresa":r["empresa_orig"],"Tipo":r["tipo_orig"],
            "ID AllRide":r["id"],"Estado":r["estado"]
        } for r in res["cancelar"]]), use_container_width=True)
        data = gen_cancelacion(res["cancelar"], f_cancel_tmpl.read())
        st.download_button(
            f"⬇️ Descargar cancelación masiva ({len(res['cancelar'])} viajes)",
            data, "Cancelacion_masiva.xlsx", use_container_width=True)

# ─── TAB 8 — TODO EN ZIP ─────────────────────────────────────────────────────
with tabs[7]:
    st.subheader("📦 Descargar paquete completo")
    st.info("Genera un ZIP con todos los archivos organizados en carpetas por paso.")

    if st.button("🚀 Generar paquete completo", type="primary", use_container_width=True):
        todos = {}

        # 1. Paradas
        if f_stops_list and f_stops_tmpl:
            adv, falt = analizar_paradas(cli, stops_df, routes_df)
            if falt:
                todos["01_paradas/paradas_a_crear.xlsx"] = gen_stops_creation(falt, f_stops_tmpl.read())

        # 2. Rutas
        if f_routes_list and f_routes_tmpl:
            rutas_n = analizar_rutas(cli, routes_df)
            if rutas_n:
                todos["02_rutas/rutas_a_crear.xlsx"] = gen_routes_creation(rutas_n, f_routes_tmpl.read())

        # 3. Edición regulares
        if editar_reg:
            tmpl = f_edit_reg_tmpl.read() if f_edit_reg_tmpl else None
            todos["03_editar_reg/Edicion_horarios_regulares.xlsx"] = gen_edicion_horarios(editar_reg, tmpl)

        # 4. Edición SPOT
        if editar_spot:
            tmpl = f_edit_spot_tmpl.read() if f_edit_spot_tmpl else None
            todos["04_editar_spot/Edicion_horarios_SPOT.xlsx"] = gen_edicion_horarios(editar_spot, tmpl)

        # 5. OTS regulares
        if crear_reg and f_ots_tmpl:
            for k,v in gen_one_time_services(crear_reg, f_ots_tmpl.read()).items():
                todos[f"05_crear_reg/{k}"] = v

        # 6. ODD SPOT
        if crear_spot and f_odd_tmpl and f_stops_list:
            for k,v in gen_odd_spots(crear_spot, f_odd_tmpl.read(), stops_df).items():
                todos[f"06_crear_spot/{k}"] = v

        # 7. Cancelación
        todos["07_cancelar/Cancelacion_masiva.xlsx"] = gen_cancelacion(res["cancelar"], f_cancel_tmpl.read())

        zip_data = gen_zip(todos)
        st.download_button("⬇️ Descargar paquete completo (.zip)", zip_data,
            "allride_cuadratura_completo.zip", mime="application/zip", use_container_width=True)
        st.success(f"✅ Paquete con {len(todos)} archivos en {len(set(k.split('/')[0] for k in todos))} carpetas.")

st.divider()
st.caption("AllRide Cuadratura v2.0 · Flujo: Paradas → Rutas → Edición → Creación → Cancelación")
